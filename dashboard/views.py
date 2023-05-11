from django.shortcuts import render,redirect
from django.http import HttpResponse
from .models import Product,Issued_Items
from django.contrib.auth.decorators import login_required
from .forms import ProductForm,orderform,sendemailform
from django.contrib.auth.models import User
from django.contrib import messages
from django.core.mail import send_mail
from django.conf import settings
from reportlab.pdfgen import canvas


from django.http import HttpResponse
from openpyxl import Workbook
from .models import Product, Issued_Items
from django.utils import timezone
from django.http import HttpResponseBadRequest
# Create your views here.

@login_required
def index(request):
    issueditems=Issued_Items.objects.all()
    product=Product.objects.all()
    items_count = issueditems.count()
    product_count = product.count()
    workers_count =User.objects.all().count()
    
    if request.method=='POST':
        emailform=sendemailform(request.POST)
        if emailform.is_valid():
            name=emailform.cleaned_data['title']
            toemail=emailform.cleaned_data['empemail']
            message=emailform.cleaned_data['message']
            send_mail(name,message,'settings.EMAIL_HOST_USER',[toemail],fail_silently=False)
            return redirect('dashboardindex')
    else:
        emailform=sendemailform()
    if request.method=='POST':
         form=orderform(request.POST)
         if form.is_valid():
            instance=form.save(commit=False)
            instance.staff=request.user
            instance.save()
            return redirect('dashboardindex')
    else:
        form=orderform()
    context={
        'issueditems':issueditems,
        'form':form,
        'product':product,
        'product_count': product_count,
        'workers_count': workers_count,
        'items_count': items_count,
        'emailform':emailform,

    }
    return render(request,'dashboard/index.html',context)
@login_required
def staff(request):
    workers=User.objects.all()
    workers_count = workers.count()
    items_count =Issued_Items.objects.all().count()
    product_count =Product.objects.all().count()
    if request.method=='POST':
        emailform=sendemailform(request.POST)
        if emailform.is_valid():
            name=emailform.cleaned_data['title']
            toemail=emailform.cleaned_data['empemail']
            message=emailform.cleaned_data['message']
            send_mail(name,message,'settings.EMAIL_HOST_USER',[toemail],fail_silently=False)
            return redirect('dashboardindex')
    else:
        emailform=sendemailform()
    context= {
        'workers':workers,
        'workers_count': workers_count,
        'items_count': items_count,
        'product_count': product_count,
        'emailform':emailform,
    }
    return render(request,'dashboard/staff.html',context)
@login_required
def staff_detail(request,pk):
    worker=User.objects.all()
    workers_count = worker.count()
    items_count =Issued_Items.objects.all().count()
    product_count =Product.objects.all().count()
    if request.method=='POST':
        emailform=sendemailform(request.POST)
        if emailform.is_valid():
            name=emailform.cleaned_data['title']
            toemail=emailform.cleaned_data['empemail']
            message=emailform.cleaned_data['message']
            send_mail(name,message,'settings.EMAIL_HOST_USER',[toemail],fail_silently=False)
            return redirect('dashboardindex')
    else:
        emailform=sendemailform()
    workers=User.objects.get(id=pk)
    context={
        'workers':workers,
        'emailform':emailform,
        'worker':worker,
        'workers_count': workers_count,
        'items_count': items_count,
        'product_count': product_count,
    }

    return render(request,'dashboard/staff_detail.html',context)
@login_required
def product(request):
    items=Product.objects.all()
    product_count = items.count()
    #items=Product.objects.raw('SELECT * FROM dashboard_product')
    workers_count =User.objects.all().count()
    items_count =Issued_Items.objects.all().count()

    if request.method=='POST':
        form=ProductForm(request.POST)
        if form.is_valid():
            form.save()
            product_name = form.cleaned_data.get('name')
            messages.success(request, f'{product_name} has been added')
            return redirect('dashboardproduct')
    else:
        form=ProductForm()
    context={
        'items':items,
        'form':form,
        'workers_count': workers_count,
        'items_count': items_count,
        'product_count': product_count,
        
    }
    return render(request,'dashboard/product.html',context)   
@login_required
def product_delete(request,pk):
    item=Product.objects.get(id=pk)
    if request.method=='POST':
        item.delete()
        return redirect('dashboardproduct')
    return render(request,'dashboard/product_delete.html') 
@login_required
def product_update(request,pk):
    item=Product.objects.get(id=pk)
    if request.method=='POST':
        form=ProductForm(request.POST,instance=item)
        if form.is_valid():
            form.save()
            return redirect('dashboardproduct')
    else:
        form=ProductForm(instance=item)
    context={
        'form':form

    }
    return render(request,'dashboard/product_update.html',context)

@login_required
def issued_items(request):
 items=Issued_Items.objects.all().order_by('-date')
 items_count = items.count()
 workers_count =User.objects.all().count()
 product_count =Product.objects.all().count()
 if request.method=='POST':
        emailform=sendemailform(request.POST)
        if emailform.is_valid():
            name=emailform.cleaned_data['title']
            toemail=emailform.cleaned_data['empemail']
            message=emailform.cleaned_data['message']
            send_mail(name,message,'settings.EMAIL_HOST_USER',[toemail],fail_silently=False)
            return redirect('dashboardindex')
        
 else:
    emailform=sendemailform()
 context={
     'items':items,
     'workers_count': workers_count,
     'items_count': items_count,
     'product_count': product_count,
     'emailform':emailform,
 }

 return render(request,'dashboard/issueditems.html',context)

def generate_pdf(request):
    # Get all products and issued items
    products = Product.objects.all()
    issued_items = Issued_Items.objects.all()

    # Create a response object with PDF content type
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="products_{timezone.now().strftime("%Y-%m-%d")}.pdf"' # Add date to filename

    # Create a canvas object and set up the font and font size
    p = canvas.Canvas(response)
    p.setFont('Helvetica', 14) # Set font and font size for the document

    # Draw the title and date
    p.drawString(200, 800, "C-DAC Inventory Report")
    p.drawString(50, 750, "List of Products")
    p.drawString(450, 750, timezone.now().strftime("%Y-%m-%d %H:%M:%S")) # Add current date and time
    p.drawString(50, 720, "--------------------------------------------------")

    # Draw the header for the products list
    p.drawString(20, 680, "Asset tag")
    p.drawString(150, 680, "Products")
    p.drawString(280, 680, "Quantity")
    p.drawString(360, 680, "Model")
    p.drawString(520, 680, "Price")
    p.drawString(20, 660, "---------------------------------------------------------------------------------------------------------------------")

    # Draw the products list
    y = 640
    for product in products:
        p.drawString(20, y, f"{product.asset}")
        p.drawString(150, y, f"{product.name}")
        p.drawString(280, y, f"{product.quantity}")
        p.drawString(360, y, f"{product.model}")
        p.drawString(520, y, f"{product.price}")
        y -= 20
    p.showPage()   

    # Draw the header for the issued items list
    p.setFont('Helvetica', 14) 
    p.drawString(50, 750, "List of Issued Items")
    p.drawString(50, 720, "--------------------------------------------------")
    p.drawString(50, 680, "Issued Items")
    p.drawString(200, 680, "Issued To")
    p.drawString(340, 680, "Location")
    p.drawString(50, 660, "----------------------------------------------------------------------------")
    

    # Draw the issued items list
    y = 640
    for issued_item in issued_items:
        p.drawString(50, y, f"{issued_item.product.name}")
        p.drawString(200, y, f"{issued_item.staff}")
        p.drawString(340, y, f"{issued_item.location}")
        
        
        y -= 20

    p.showPage()

    p.save()
    return response


def generate_excel_file(request):
    # Create a new workbook and add a worksheet
    wb = Workbook()
    ws_products = wb.active
    ws_products.title = "Products"
    ws_issued_items = wb.create_sheet("Issued Items")

    # Write headers to worksheet for Products
    ws_products['A1'] = 'Asset'
    ws_products['B1'] = 'SNO'
    ws_products['C1'] = 'Name'
    ws_products['D1'] = 'Category'
    ws_products['E1'] = 'Quantity'
    ws_products['F1'] = 'Model'
    ws_products['G1'] = 'Price'

    # Write headers to worksheet for Issued_Items
    ws_issued_items['A1'] = 'Product'
    ws_issued_items['B1'] = 'Issued to'
    ws_issued_items['C1'] = 'Quantity'
    ws_issued_items['D1'] = 'Location'

    # Write data to Products worksheet
    products = Product.objects.all()
    for row, product in enumerate(products, start=2):
        
        ws_products.cell(row=row, column=1, value=product.asset)
        ws_products.cell(row=row, column=2, value=product.sno)
        ws_products.cell(row=row, column=3, value=product.name)
        ws_products.cell(row=row, column=4, value=product.category)
        ws_products.cell(row=row, column=5, value=product.quantity)
        ws_products.cell(row=row, column=6, value=product.model)
        ws_products.cell(row=row, column=7, value=product.price)

    # Write data to Issued_Items worksheet
    issued_items = Issued_Items.objects.all()
    for row, issued_item in enumerate(issued_items, start=2):
        
        ws_issued_items.cell(row=row, column=1, value=issued_item.product.name)
        ws_issued_items.cell(row=row, column=2, value=issued_item.staff.username)
        ws_issued_items.cell(row=row, column=3, value=issued_item.issueditem_quantity)
        ws_issued_items.cell(row=row, column=4, value=issued_item.location)

    # Set the response headers
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=products.xlsx'

    # Save the workbook to the response
    wb.save(response)
    
    

    return response

def item_review(request,pk):
    items=Issued_Items.objects.all()
    item=Issued_Items.objects.get(id=pk)
    products=item.product
    items_count = items.count()
    workers_count =User.objects.all().count()
    product_count =Product.objects.all().count()
    if request.method=='POST':
        emailform=sendemailform(request.POST)
        if emailform.is_valid():
            name=emailform.cleaned_data['title']
            toemail=emailform.cleaned_data['empemail']
            message=emailform.cleaned_data['message']
            send_mail(name,message,'settings.EMAIL_HOST_USER',[toemail],fail_silently=False)
            return redirect('dashboardindex')
        status=request.POST.get('status')
        if products.quantity <= 0:
         
         if status=='Accepted':
          item.status='Pending'
          item.is_accepted=False
          item.save()
          return HttpResponseBadRequest("error:Product is out of stock!! Please order "+ products.name )
         else:
             if status=='Rejected' and item.status=='Accepted':
              products.quantity+=item.issueditem_quantity
              products.save()
             item.status='Rejected'
             item.is_accepted=False
             item.save()
        else:
         if status=='Accepted'and item.status=='Pending':
               products.quantity -= item.issueditem_quantity
               products.save()
               item.is_accepted=True
               item.status=status
               item.save()
               
         elif status=='Rejected' and item.status=='Accepted':
            products.quantity+=item.issueditem_quantity
            products.save()
            item.status=status
            item.is_accepted=False
            item.save()
         elif status=='Accepted'and item.status=='Rejected':
               products.quantity -= item.issueditem_quantity
               products.save()
               item.is_accepted=True
               item.status=status
               item.save()
               
         else:
             item.status=status
             item.save()

              
        if status == 'Accepted':   
         return redirect('dashboardissueditems')
        else:
            return redirect('dashboardissueditems')
            
        
    else:
     emailform=sendemailform()
    context={
     'items':items,
     'workers_count': workers_count,
     'items_count': items_count,
     'product_count': product_count,
     'emailform':emailform,
     'item':item
    }
    return render(request,'dashboard/item_review.html',context)